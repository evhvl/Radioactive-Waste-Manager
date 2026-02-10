import os, sqlite3
from datetime import datetime, timedelta
from tkinter import messagebox, Frame, Label, Button, filedialog, ttk
from openpyxl import Workbook, load_workbook
from constants import *


DISPOSALS_DIR = "Disposals"
REGISTRY_DB = os.path.join(DISPOSALS_DIR, "registry.sqlite")
DATE_FORMAT = "%d-%m-%Y"
HOUR_FORMAT = "%H:%M"

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

#=====SQLITE CREATE TABLES=====
def init_registry():
    ensure_dir(DISPOSALS_DIR)
    conn = sqlite3.connect(REGISTRY_DB)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY, value TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS batches (id INTEGER PRIMARY KEY AUTOINCREMENT, folder_path TEXT, created_at TEXT, finalized_at TEXT)""")
    conn.commit()
    conn.close()

#=====CREATE NEW BATCH=====
def create_new_batch_folder():
    init_registry()
    year = datetime.now().strftime("%Y")
    creation_date = datetime.now().strftime(DATE_FORMAT)
    year_dir = os.path.join(DISPOSALS_DIR, year)
    ensure_dir(year_dir)
    batch_path = os.path.join(year_dir, f"Batch__{creation_date}")
    if not os.path.exists(batch_path):
        ensure_dir(batch_path)
    conn = sqlite3.connect(REGISTRY_DB)
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM batches WHERE folder_path=?", (batch_path,))
    exists = cur.fetchone()
    if not exists:
        cur.execute("INSERT INTO batches (folder_path, created_at, finalized_at) VALUES (?,?,NULL)", (batch_path, creation_date))
    cur.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('active_batch',?)", (batch_path,))
    conn.commit()
    conn.close()
    return batch_path

#=====GET ACTIVE BATCH (IF IT DOESN'T EXIST CREATE NEW)=====
def get_active_batch():
    init_registry()
    conn = sqlite3.connect(REGISTRY_DB)
    cur = conn.cursor()
    row = cur.execute("SELECT value FROM settings WHERE key='active_batch'").fetchone()
    conn.close()
    if row and os.path.isdir(row[0]):
        return row[0]
    return create_new_batch_folder()

#=====FINALIZE CURRENT ACTIVE BATCH AND START NEW=====
def finalize_active_batch():
    init_registry()
    old_batch = get_active_batch()
    finalized_date = datetime.now().strftime(DATE_FORMAT)
    conn = sqlite3.connect(REGISTRY_DB)
    cur = conn.cursor()
    cur.execute("UPDATE batches SET finalized_at=? WHERE folder_path=? AND finalized_at IS NULL", (finalized_date, old_batch))
    conn.commit()
    conn.close()
    new_batch = create_new_batch_folder()
    return old_batch, new_batch

#=====CREATE SQLITE+XLSX FILES INSIDE BATCH FOLDER=====
def init_storage_files(batch_path):
    db_path = os.path.join(batch_path, "storage.sqlite")
    xlsx_path = os.path.join(batch_path, "storage.xlsx")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS stored_vials (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                            radionuclide TEXT,
                                                            source_db TEXT,
                                                            calibration_date TEXT,
                                                            stored_at TEXT,
                                                            activity_mci REAL,
                                                            permitted_date TEXT,
                                                            recommended_date TEXT,
                                                            limit_bq REAL,
                                                            limit_mci REAL)""")
    conn.commit()
    conn.close()
    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stored Vials"
        ws.append(["ID", "Radionuclide", "Source DB", "Calibration Date", "Stored at", "Activity(mCi)", "Permitted Date", "Recommended Date", "Limit (mCi)", "Limit (Bq)"])
        wb.save(xlsx_path)
    return db_path, xlsx_path

#=====DECAY AND DATES CALCULATIONS=====
def bq_to_mci(bq):
    return float(bq) / 3.7e7

def mci_to_bq(mci):
    return float(mci) * 3.7e7

def decay_activity(activity_mci, half_life_hours, delta_hours):
    lambda_ = math.log(2) / half_life_hours
    return float(activity_mci) * math.exp(-lambda_ * float(delta_hours))

def calc_date_below_limit(activity_mci, half_life_hours, limit_bq, start_date):
    activity_mci = float(activity_mci)
    half_life_hours = float(half_life_hours)
    if half_life_hours <= 0:
        return start_date.strftime(DATE_FORMAT)
    if limit_bq is None:
        return start_date.strftime(DATE_FORMAT)
    limit_mci = round(float(limit_bq) / 3.7e7, 2)
    if limit_mci <= 0:
        return start_date.strftime(DATE_FORMAT)
    if float(activity_mci) <= limit_mci:
        return start_date.strftime(DATE_FORMAT)
    lambda_ = math.log(2) / half_life_hours
    ratio = activity_mci / limit_mci
    if ratio <= 1:
        return start_date.strftime(DATE_FORMAT)
    t_hours = math.log(ratio) / lambda_
    return (start_date + timedelta(hours=t_hours)).strftime(DATE_FORMAT)

def calc_recommended_and_permitted_date(radionuclide, activity_mci, stored_at, safety_factor=0.1):
    if radionuclide not in DISPOSAL_LIMITS_BQ:
        messagebox.showerror("Missing disposal limit", f"No disposal limit defined for {radionuclide}.")
        return None, None, None
    start_date = datetime.strptime(stored_at, DATE_FORMAT)
    half_life = next(hl for name, hl in VIAL_DATA if name == radionuclide)
    limit_bq = DISPOSAL_LIMITS_BQ[radionuclide]
    permitted = calc_date_below_limit(activity_mci, half_life, limit_bq, start_date)
    recommended = calc_date_below_limit(activity_mci, half_life, limit_bq * safety_factor, start_date)
    return recommended, permitted, float(limit_bq)

#=====STORE VIAL IN SQLITE+XLSX=====
def store_vial(radionuclide, source_db, calibration_date, stored_at, activity_mci, permitted_date=None, recommended_date=None, limit_bq=None):
    if permitted_date is None or recommended_date is None or limit_bq is None:
        recommended_date, permitted_date, limit_bq = calc_recommended_and_permitted_date(radionuclide, float(activity_mci), stored_at)
    limit_mci = round(bq_to_mci(limit_bq), 2)
    batch_path = get_active_batch()
    db_path, xlsx_path = init_storage_files(batch_path)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("INSERT INTO stored_vials (radionuclide, source_db, calibration_date, stored_at, activity_mci, permitted_date, recommended_date, limit_bq, limit_mci) VALUES (?,?,?,?,?,?,?,?,?)",
                (radionuclide, source_db, calibration_date, stored_at, float(activity_mci), permitted_date, recommended_date, float(limit_bq), float(limit_mci)))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    wb = load_workbook(xlsx_path)
    ws = wb["Stored Vials"]
    ws.append([new_id, radionuclide, source_db, calibration_date, stored_at, float(activity_mci), permitted_date, recommended_date, float(limit_mci), float(limit_bq)])
    wb.save(xlsx_path)
    return new_id, batch_path

#=====READ STORED VIALS FROM BATCH=====
def read_stored_vials(batch_path=None):
    if batch_path is None:
        batch_path = get_active_batch()
    db_path = os.path.join(batch_path, "storage.sqlite")
    if not os.path.exists(db_path):
        return [], batch_path
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    rows = cur.execute("SELECT ID, radionuclide, stored_at, activity_mci, permitted_date, recommended_date, limit_mci FROM stored_vials ORDER BY id").fetchall()
    conn.close()
    return rows, batch_path

#=====TREE STATUS=====
def disposal_status(recommended_date, permitted_date):
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    rec = datetime.strptime(recommended_date, DATE_FORMAT)
    perm = datetime.strptime(permitted_date, DATE_FORMAT)
    if today < rec:
        return "OK"
    elif today < perm:
        return "WAIT"
    return "READY"

#=====DISPOSAL SUMMARY=====
def disposal_summary(rows):
    total_vials = len(rows)
    total_activity = sum(float(r[3]) for r in rows)
    ready_count = 0
    for r in rows:
        permitted = r[4]
        recommended = r[5]
        if disposal_status(recommended, permitted) == "READY":
            ready_count += 1
    return total_vials, round(total_activity, 2), ready_count

#=====DISPOSAL TAB=====
def build_disposal_tab(parent_tab, *, on_back=None):
    for w in parent_tab.winfo_children():
        w.destroy()
    state = {"batch_path": None, "read_only": False}
    header = Label(parent_tab, text="DISPOSAL STORAGE", **TEXT_COLORS, font=(FONT_NAME, 18, "bold"))
    header.pack(pady=(10,5), fill="x")
    frame = Frame(parent_tab, bg=C4)
    frame.pack(pady=5)
    batch_label = Label(frame, text="", **TEXT_COLORS, font=(FONT_NAME, 12, "bold"))
    batch_label.grid(row=0, column=0, padx=10)
    mode_label = Label(frame, text="", bg=C4, fg="orange", font=(FONT_NAME, 10, "bold"))
    mode_label.grid(row=1, column=0, padx=10)
    #Fuctions
    def load_batch(batch_path, read_only=False):
        state["batch_path"] = batch_path
        state["read_only"] = read_only
        rows, _ = read_stored_vials(batch_path)
        batch_label.config(text=f"Viewing: {os.path.basename(batch_path)}")
        if read_only:
            mode_label.config(text="READ ONLY (Finalized/Old Batch)")
        else:
            mode_label.config(text="ACTIVE BATCH")
        tree.delete(*tree.get_children())
        for r in rows:
            internal_id = r[0]
            visible = r[1:]
            status = disposal_status(visible[4], visible[3])
            tree.insert("", "end", iid=internal_id, values=list(visible) + [status], tags=(status,))
            total_vials, total_activity, ready_count = disposal_summary(rows)
            summary_label.config(text=f"Total vials: {total_vials}   |   Total activity: {total_activity} mCi   |   READY: {ready_count}")
        if read_only:
            finalize_btn.config(state="disabled")
        else:
            finalize_btn.config(state="normal")
    def refresh():
        if not state["batch_path"]:
            load_active()
            return
        load_batch(state["batch_path"], read_only=state["read_only"])
    def load_active():
        active = get_active_batch()
        load_batch(active, read_only=False)
    def open_old_batch():
        folder = filedialog.askdirectory(title="Select Old (Finalized) Batch Folder", initialdir=DISPOSALS_DIR)
        if not folder:
            return
        load_batch(folder, read_only=True)
    def finalize_batch():
        if state["read_only"]:
            return
        if not messagebox.askyesno("Finalize Batch", "Are you sure you want to finalize this ACTIVE batch?\nAfter finalize, new stored vials will go into a NEW batch."):
            return
        old_batch, new_batch = finalize_active_batch()
        messagebox.showinfo("Batch Finalized", f"Old batch closed:\n{os.path.basename(old_batch)}\n\nNew active batch:\n{os.path.basename(new_batch)}")
        load_batch(new_batch, read_only=False)
    btns = Frame(parent_tab, bg=C4)
    btns.pack(pady=(10,15))
    Button(btns, text="Refresh", **{k: v for k, v in TAB_BUTTON_STYLE.items() if k not in ['width', 'height', 'font']}, width=10, height=1, font=(FONT_NAME, 10, "bold"), command=refresh).grid(row=0, column=0, padx=6)
    Button(btns, text="Open Old Batch",**{k: v for k, v in TAB_BUTTON_STYLE.items() if k not in ['width', 'height', 'font']}, width=14, height=1, font=(FONT_NAME, 10, "bold"), command=open_old_batch).grid(row=0, column=1, padx=6)
    Button(btns, text="Back to Active", **{k: v for k, v in TAB_BUTTON_STYLE.items() if k not in ['width', 'height', 'font']}, width=14, height=1, font=(FONT_NAME, 10, "bold"), command=load_active).grid(row=0, column=2, padx=6)
    finalize_btn = Button(btns, text="✗Finalize Batch✗", **{k: v for k, v in TAB_BUTTON_STYLE.items() if k not in ['width', 'height', 'font']}, width=16, height=1, font=(FONT_NAME, 10, "bold"), command=finalize_batch)
    finalize_btn.grid(row=0, column=3, padx=6)
    #Tree
    columns = [("Vial", 80), ("Stored at", 95), ("Activity (mCi)", 110), ("Permitted Date", 120), ("Recommended Date", 150), ("Limit (mCi)", 105), ("Status", 85),]
    tree = ttk.Treeview(parent_tab, columns=[c[0] for c in columns], show="headings", height=12)
    tree.pack(pady=(10,0))
    for col_name, col_width in columns:
        tree.heading(col_name, text=col_name)
        tree.column(col_name, anchor="center", width=col_width, stretch=False)
    style = ttk.Style()
    style.theme_use("default")
    style.configure("Treeview", background=C2, fieldbackground=C2, foreground="black", rowheight=26, borderwidth=1, bordercolor="black", relief="solid")
    style.configure("Treeview.Heading", background=C3, foreground="white", font=(FONT_NAME, 11, "bold"), relief="solid")
    style.map("Treeview", background=[("selected", "#8FAADC"), ("!selected", C2)], foreground=[("selected", "black")])
    style.layout("Treeview", [("Treeview.treearea", {"sticky": "nsew"})])
    tree.tag_configure("OK", background="#CDECCF")
    tree.tag_configure("WAIT", background="#FFE9B3")
    tree.tag_configure("READY", background="#F6B3B3")
    summary_frame = Frame(parent_tab, bg=C4)
    summary_frame.pack(pady=(5,10))
    summary_label = Label(summary_frame, text="", **TEXT_COLORS, font=(FONT_NAME, 12, "bold"))
    summary_label.pack()
    if on_back:
        Button(parent_tab, text="Back", **TAB_BUTTON_STYLE, command=on_back).pack(pady=(0,10))
    load_active()