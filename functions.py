from datetime import datetime, timedelta
from typing import Optional
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from constants import *
from tkinter import Frame, Scrollbar, Canvas, messagebox, Button, END
import os, sqlite3
from openpyxl import Workbook, load_workbook

#Center Window According to Screen
def center_window(window, w, h):
    window.resizable(False, False)
    window.update_idletasks()
    screen_w = window.winfo_screenwidth()
    screen_h = window.winfo_screenheight()
    x = (screen_w // 2) - (w // 2)
    y = (screen_h // 2) - (h // 2)
    window.geometry(f"{w}x{h}+{x}+{y}")
    window.grab_set()

#Create Scrollable Frame
def create_scrollable_frame(parent):
    contents = Frame(parent, bg=C4)
    contents.pack(fill="both", expand=True, pady=(5, 0))
    canvas = Canvas(contents, bg=C4, highlightthickness=0)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar = Scrollbar(contents, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)
    scroll_frame = Frame(canvas, bg=C4)
    scroll_window = canvas.create_window((0, 0), window=scroll_frame, anchor="n")
    def update_scroll_region(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(scroll_window, width=canvas.winfo_width())
    scroll_frame.bind("<Configure>", update_scroll_region)
    def _on_mousewheel(event):
        canvas.yview_scroll(-1 * int(event.delta / 120), "units")
    canvas.bind_all("<MouseWheel>", _on_mousewheel)
    return contents, canvas, scroll_frame, scrollbar

#Update Time for Entry
def update_time(time_entry):
    now = datetime.now().strftime("%H:%M")
    time_entry.delete(0, END)
    time_entry.insert(0, now)

#Create Excel for Vials
def create_excel_for_vial(excel_path):
    if os.path.exists(excel_path):
        return
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Vial Info"
    ws_info.append(["Batch Number", "Serial Number", "Calibration Date", "Calibration Time", "Activity(mCi)", "Volume(ml)", "Concentration(mCi/ml)", "Expiration Date", "Stored Date", "Disposal Date"])
    ws_admin = wb.create_sheet("Administrations")
    ws_admin.append(["ID", "Date", "Time", "Patient Name", "Concentration(mCi/ml)", "Dose(mCi)", "Volume(ml)", "Volume Left(ml)"])
    wb.save(excel_path)

#Create Excel for Tc99m Gen
def create_excel_for_tc99m(excel_path):
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Gen Info"
        ws.append(["Gen ID", "Calibration Date", "Calibration Time", "Mo99 Activity (mCi)", "Delivery Date", "Expiration Date", "Stored Date", "Disposal Date"])
        ws2 = wb.create_sheet("Elutions")
        ws2.append(["", "Date", "Time", "Activity(mCi)", "Expected(mCi)", "Div(%)", "Vol(ml)", "Conc(mCi/ml)"])
        ws3 = wb.create_sheet("Kits")
        ws3.append(["Kit ID", "Patient ID", "Date", "Time", "Kit", "Volume", "Activity(mCi)", "Conc(mCi/ml)", "Dose(mCi)", "Dose Volume(ml)", "Volume Left(ml)", "Patient Name"])
        wb.save(excel_path)

#Create Excel for Ga68 Gen
def create_excel_for_ga68(excel_path):
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Gen Info"
        ws.append(["Gen ID", "Model", "Delivery Date", "Calibration Date", "Calibration Time", "Activity (MBq)", "Expiration Date", "Disposal Date"])
        ws2 = wb.create_sheet("Elutions")
        ws2.append(["", "Date", "Time", "Activity(mCi)"])
        ws3 = wb.create_sheet("DOTATOC")
        ws3.append(["", "Date", "Patient", "Weight (kg)", "Admin Time", "Dose (mCi)", "Volume (ml)", "Concentration (mCi/ml)", "ITLC(<2%)", "Residual(mCi)"])
        wb.save(excel_path)

#Excel Helpers for Ga68
def get_dotatoc_excel_path(dbfile):
    folder = os.path.dirname(dbfile)
    return os.path.join(folder, f"{os.path.basename(folder)}.xlsx")
def update_dotatoc_excel(dbfile, row_id, new_dose, new_volume, new_concentration, new_itlc, new_residual):
    excel_path = get_dotatoc_excel_path(dbfile)
    wb = load_workbook(excel_path)
    ws = wb["DOTATOC"]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).strip() == str(row_id).strip():
            ws.cell(row=r, column=6, value=new_dose)
            ws.cell(row=r, column=7, value=new_volume)
            ws.cell(row=r, column=8, value=new_concentration)
            ws.cell(row=r, column=9, value=new_itlc)
            ws.cell(row=r, column=10, value=new_residual)
            break
    wb.save(excel_path)

#Add New Row to Excel Sheets
def append_row_to_sheet(excel_path, sheet_name, row_values):
   if not os.path.exists(excel_path):
       if "tc" in excel_path.lower():
           create_excel_for_tc99m(excel_path)
       elif "ga" in excel_path.lower():
           create_excel_for_ga68(excel_path)
       elif "vial" in excel_path.lower():
           create_excel_for_vial(excel_path)
       else:
           raise ValueError(f"Cannot determine generator type from path: {excel_path}.")
   wb = load_workbook(excel_path)
   if sheet_name not in wb.sheetnames:
       ws = wb.create_sheet(sheet_name)
   else:
       ws = wb[sheet_name]
   ws.append(row_values)
   wb.save(excel_path)

#Find the Patient in Excel and Insert New Data
def find_patient_insert_row(ws, parent_id):
    last_child_row = None
    parent_row = None
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val is None:
            continue
        cell_val = str(cell_val)
        if cell_val == str(parent_id):
            parent_row = r
        elif cell_val.startswith(f"{parent_id}."):
            last_child_row = r
    if last_child_row:
        return last_child_row + 1
    elif parent_row:
        return parent_row + 1
    else:
        return ws.max_row + 1

#Renumber Child Rows After Delete
def renumber_children(conn, ws, tree, parent_iid):
    cur = conn.cursor()
    cur.execute("""SELECT id FROM kits WHERE parent_id=? ORDER BY time""", (parent_iid,))
    children = [row[0] for row in cur.fetchall()]
    max_idx = 0
    for cid in children:
        parts = cid.split(".")
        if len(parts) == 2 and parts[0] == str(parent_iid):
            try:
                seq = int(parts[1])
                if seq > max_idx:
                    max_idx = seq
            except ValueError:
                continue
    next_idx = max_idx + 1
    for old_id in children:
        pass
    for r in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val and str(cell_val).startswith(f"{parent_iid}."):
            ws.cell(row=r, column=1).value = cell_val
            ws.parent_id = parent_iid
    children_iids = list(tree.get_children(parent_iid))
    for old_iid in children_iids:
        new_iid = old_iid
        values = tree.item(old_iid, "values")
        index = tree.index(old_iid)
        tree.delete(old_iid)
        tree.insert(parent_iid, index, iid=new_iid, values=values)
    conn.commit()
    return next_idx

#Find Last Folder When Opening Old File
def find_last_folder(base_dir, subfolder=None):
    root = os.path.join(base_dir, subfolder) if subfolder else base_dir
    os.makedirs(root, exist_ok=True)
    now = datetime.now()
    preferred = os.path.join(root, now.strftime("%Y"), now.strftime("%m"))
    if os.path.exists(preferred):
        return preferred
    best_folder = None
    best_distance = None
    for year_name in os.listdir(root):
        year_path = os.path.join(root, year_name)
        if not os.path.isdir(year_path):
            continue
        for month_name in os.listdir(year_path):
            month_path = os.path.join(year_path, month_name)
            if not os.path.isdir(month_path):
                continue
            try:
                folder_date = datetime.strptime(f"{year_name}-{month_name}", "%Y-%m")
            except ValueError:
                continue
            distance = abs((folder_date - now).days)
            if best_distance is None or distance < best_distance:
                best_distance = distance
                best_folder = month_path
    return best_folder if best_folder else root

#Store Gens
def store_gen(*, conn, dbfile, excel_sheet="Gen Info", date_format=DATE_FORMAT, on_store_callback=None):
    if not messagebox.askyesno("Store Generator", "Are you sure you want to store this generator?\nThis action cannot be undone."):
        return False
    stored_date = datetime.now().strftime(date_format)
    cur = conn.cursor()
    cur.execute("UPDATE generator_info SET stored_date=?", (stored_date,))
    conn.commit()
    folder = os.path.dirname(dbfile)
    excel_path = os.path.join(folder, f"{os.path.basename(folder)}.xlsx")
    wb = load_workbook(excel_path)
    ws = wb[excel_sheet]
    ws.cell(row=2, column=7).value = stored_date
    wb.save(excel_path)
    wb.close()
    messagebox.showinfo("Stored", f"Generator stored on {stored_date}.")
    if on_store_callback:
        on_store_callback()
    return True

#Dispose Gens
def dispose_gen(*, conn, dbfile, excel_sheet="Gen Info", date_format="%d-%m-%Y", on_disposed_callback=None):
    if not messagebox.askyesno("Dispose Generator", "Are you sure you want to dispose this generator?\nThis action cannot be undone."):
        return False
    disposal_date = datetime.now().strftime(date_format)
    cur = conn.cursor()
    cur.execute("UPDATE generator_info SET disposal_date=?", (disposal_date,))
    conn.commit()
    folder = os.path.dirname(dbfile)
    excel_path = os.path.join(folder, f"{os.path.basename(folder)}.xlsx")
    wb = load_workbook(excel_path)
    ws = wb[excel_sheet]
    ws.cell(row=2, column=8).value = disposal_date
    wb.save(excel_path)
    wb.close()
    messagebox.showinfo("Disposed", f"Generator disposed on {disposal_date}.")
    if on_disposed_callback:
        on_disposed_callback()
    return True

#Disable Buttons after Stored or Expired
def disable_buttons(parent, exempt_texts=None):
    if exempt_texts is None:
        exempt_texts = []
    for widget in parent.winfo_children():
        if isinstance(widget, Button):
            if widget.cget("text") not in exempt_texts:
                widget.config(state="disabled")
        elif widget.winfo_children():
            disable_buttons(widget, exempt_texts)

#Update Headers and Disable Buttons after Stored or Expired
def update_header_and_disable(cur, header, tab, is_stored=False, is_disposed=False, is_expired=False):
    if is_disposed:
        cur.execute("SELECT disposal_date FROM generator_info ORDER BY rowid DESC LIMIT 1")
        row = cur.fetchone()
        disposal_date = row[0] if row and row[0] else ""
        header.config(text=f"⚠ GENERATOR DISPOSED ({disposal_date}) – NO FURTHER ACTIONS ALLOWED", fg="#660000", highlightthickness=0, font=(FONT_NAME,23,"bold"))
        disable_buttons(tab, exempt_texts=["Back", "Load"])
    elif is_stored and not is_disposed:
        cur.execute("SELECT stored_date FROM generator_info ORDER BY rowid DESC LIMIT 1")
        row = cur.fetchone()
        stored_date = row[0] if row and row[0] else ""
        header.config(text=f"⚠ GENERATOR STORED ({stored_date}) – NO FURTHER ACTIONS ALLOWED", fg="#CC0000", highlightthickness=0, font=(FONT_NAME,23,"bold"))
        disable_buttons(tab, exempt_texts=["Back", "Load", "✗Dispose Gen✗"])
    elif is_expired and not is_disposed and not is_stored:
        header.config(text="⚠ GENERATOR EXPIRED – NO FURTHER ACTIONS ALLOWED", fg="#FF8000", highlightthickness=0, font=(FONT_NAME,23,"bold"))
        disable_buttons(tab, exempt_texts=["Back", "Load", "✗Store Gen✗"])

def get_float(value, default=None):
    try:
        if hasattr(value, "get"):
            value = value.get()
        text = str(value).strip().replace(",", ".")
        return float(text)
    except (ValueError, TypeError):
        return default

#=====ADD BATCH AND SERIAL COLS IN SQL=====
def ensure_vial_info_columns(conn):
    cur = conn.cursor()
    cols = [r[1] for r in cur.execute("PRAGMA table_info(vial_info)").fetchall()]
    if "batch_number" not in cols:
        cur.execute("ALTER TABLE vial_info ADD COLUMN batch_number TEXT")
    if "serial_number" not in cols:
        cur.execute("ALTER TABLE vial_info ADD COLUMN serial_number TEXT")
    conn.commit()

#=====MARK FOLDERS AS STORED/DISPOSED=====
def update_folder_status(path, *, stored=False, disposed=False):
    try:
        if not os.path.exists(path):
            return path
        parent = os.path.dirname(path)
        current_name = os.path.basename(path)
        clean_name = current_name.replace("-ACTIVE", "").replace("-STORED", "").replace("-DISPOSED", "")
        if disposed:
            new_name = f"{clean_name}-DISPOSED"
        elif stored:
            new_name = f"{clean_name}-STORED"
        else:
            new_name = f"{clean_name}-ACTIVE"
        new_path = os.path.join(parent, new_name)
        if path != new_path and not os.path.exists(new_path):
            os.rename(path, new_path)
        return new_path
    except Exception as e:
        messagebox.showerror("Error", f"Folder rename error: {e}")
        return path

#=====PDF HELPER=====
def get_vial_batch_serial(source_db):
    try:
        conn = sqlite3.connect(source_db)
        cur = conn.cursor()
        row = cur.execute("SELECT batch_number, serial_number FROM vial_info ORDER BY rowid DESC LIMIT 1").fetchone()
        conn.close()
        if row:
            return row[0], row[1]
    except Exception:
        pass
    return "", ""

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

#=====DECAY AND DATES CALCULATIONS=====
def bq_to_mci(bq):
    return float(bq) / 3.7e7

def mci_to_bq(mci):
    return float(mci) * 3.7e7

def mci_to_kbq(mci):
    return float(mci) * 37000.0

def kbq_to_mci(kbq):
    return float(kbq) / 37000.0

def activity_conc_kbq_per_kg(activity_mci, mass_kg):
    mass_kg = float(mass_kg)
    if mass_kg <= 0:
        return messagebox.showerror("Error", "Mass cannot be <= 0.")
    return mci_to_kbq(activity_mci) / mass_kg

def disposal_fraction(activity_mci, mass_kg, limit_kbq_kg):
    conc = activity_conc_kbq_per_kg(activity_mci, mass_kg)
    if conc is None or limit_kbq_kg in (None, 0):
        return messagebox.showerror("Error", "")
    return conc / float(limit_kbq_kg)

def decay_activity(activity_mci, half_life_hours, delta_hours):
    lambda_ = math.log(2) / half_life_hours
    return float(activity_mci) * math.exp(-lambda_ * float(delta_hours))

def get_half_life_hours(radionuclide):
    if radionuclide == TC99M_NUCLIDE:
        return T12_TC99M
    if radionuclide in ("Ga-68", "Ga68"):
        return T12_GA68 / 60.0
    if radionuclide in ("Lu-177", "Lu177"):
        return T12_LU177
    return next(hl for name, hl in VIAL_DATA if name == radionuclide)

def activity_now(radionuclide, stored_at_str, activity0):
    now_dt = datetime.now()
    stored_dt = datetime.strptime(stored_at_str, DATE_FORMAT)
    delta_h = (now_dt - stored_dt).total_seconds() / 3600.0
    if delta_h < 0:
        delta_h = 0.0
    half_life = get_half_life_hours(radionuclide)
    return float(decay_activity(float(activity0), float(half_life), float(delta_h)))

def calc_date_below_limit(activity_mci, half_life_hours, limit_bq, start_date):
    activity_mci = float(activity_mci)
    half_life_hours = float(half_life_hours)
    if half_life_hours <= 0 or limit_bq in (None,0,""):
        return start_date.strftime(DATE_FORMAT)
    limit_mci = float(limit_bq) / 3.7e7
    if limit_mci <= 0 or activity_mci <= limit_mci:
        return start_date.strftime(DATE_FORMAT)
    lambda_ = math.log(2) / half_life_hours
    t_hours = math.log(activity_mci / limit_mci) / lambda_
    return (start_date + timedelta(hours=t_hours)).strftime(DATE_FORMAT)

def calc_recommended_and_permitted_date(radionuclide, activity_mci, stored_at, safety_factor=0.1):
    start_date = datetime.strptime(stored_at, DATE_FORMAT)
    half_life = get_half_life_hours(radionuclide)
    limit_bq = DISPOSAL_LIMITS_BQ.get(radionuclide)
    if limit_bq in (None,0):
        ready_date = (start_date + timedelta(hours=float(half_life) * 10.0)).strftime(DATE_FORMAT)
        return  ready_date, ready_date, None
    permitted = calc_date_below_limit(activity_mci, half_life, limit_bq, start_date)
    recommended = calc_date_below_limit(activity_mci, half_life, limit_bq * safety_factor, start_date)
    return recommended, permitted, float(limit_bq)

def calc_bag_clearance(ready_items, mass_kg):
    mass_kg = float(mass_kg)
    if mass_kg <= 0:
        return messagebox.showerror("Error", "Bag Mass must be greater than 0.")
    grouped = {}
    for it in ready_items:
        nuclide = it["radionuclide"]
        grouped.setdefault(nuclide, {"activity_now_mci": 0.0, "items": []})
        grouped[nuclide]["activity_now_mci"] += float(it["activity_now"])
        grouped[nuclide]["items"].append(it)
    details = []
    total_fraction = 0.0
    missing_limits = []
    for nuclide, g in sorted(grouped.items(), key=lambda kv: kv[0]):
        limit_kbq_kg = DISPOSAL_LIMITS_KBQ_PER_KG[nuclide]
        if limit_kbq_kg in (None,0,""):
            missing_limits.append(nuclide)
            continue
        total_activity_mci = float(g["activity_now_mci"])
        total_activity_kbq = mci_to_kbq(total_activity_mci)
        concentration = total_activity_kbq / mass_kg
        fraction = concentration / float(limit_kbq_kg)
        total_fraction += fraction
        details.append({"radionuclide": nuclide,
                        "activity_now_mci": total_activity_mci,
                        "activity_now_kbq": total_activity_kbq,
                        "mass_kg": mass_kg,
                        "concentration_kbq_kg": concentration,
                        "limit_kbq_kg": float(limit_kbq_kg),
                        "fraction": fraction,
                        "count": len(g["items"]),
                        "items": g["items"],})
    if missing_limits:
        messagebox.showwarning("Missing Limits", "Missing Table A kBq/kg limits for " + ", ".join(missing_limits))
    return details, total_fraction, total_fraction < 1.0

#=====TREE STATUS=====
def disposal_status(recommended_date, permitted_date):
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    rec = datetime.strptime(recommended_date, DATE_FORMAT)
    perm = datetime.strptime(permitted_date, DATE_FORMAT)
    if today < perm:
        return "STORED"
    elif perm <= today < rec:
        return "WAIT"
    return "READY"

#=====DISPOSAL SUMMARY=====
def disposal_summary(rows):
    total_vials = len(rows)
    now_dt = datetime.now()
    total_activity_now = 0.0
    ready_count = 0
    for r in rows:
        radionuclide = r[1]
        stored_at_str = r[3]
        activity0 = float(r[4])
        if radionuclide == TC99M_NUCLIDE:
            half_life = T12_TC99M
        else:
            half_life = next(hl for name, hl in VIAL_DATA if name == radionuclide)
        stored_dt = datetime.strptime(stored_at_str, DATE_FORMAT)
        delta_hours = (now_dt - stored_dt).total_seconds() / 3600
        if delta_hours < 0:
            delta_hours = 0.0
        activity_now = decay_activity(activity0, half_life, delta_hours)
        total_activity_now += activity_now
        permitted = r[5]
        recommended = r[6]
        if disposal_status(recommended, permitted) == "READY":
            ready_count += 1
    return total_vials, round(total_activity_now, 6), ready_count

#=====DAILY LOG EXCEL+SQLite (COMMON)=====
def get_disposed_by_date_dir(for_date: Optional[datetime] = None) -> str:
    if for_date is None:
        for_date = datetime.now()
    year = for_date.strftime("%Y")
    month = for_date.strftime("%m")
    path = os.path.join(DAILY_DISPOSALS_DIR, year, month)
    ensure_dir(path)
    return path

def get_daily_disposal_excel_path(disposal_date_str: str) -> str:
    dt = datetime.strptime(disposal_date_str, DATE_FORMAT)
    folder = get_disposed_by_date_dir(dt)
    return os.path.join(folder, f"disposal__{disposal_date_str}.xlsx")

def get_ready_vials_pdf_path(pdf_date_str: str) -> str:
    dt = datetime.strptime(pdf_date_str, DATE_FORMAT)
    folder = get_disposed_by_date_dir(dt)
    return os.path.join(folder, f"ready_vials__{pdf_date_str}.pdf")

def ensure_daily_log_workbook(xlsx_path: str):
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        ws0 = wb.active
        wb.remove(ws0)
    vial_headers = ["Disposal Date", "Disposal Time", "Radionuclide", "Calibration Date", "Stored At", "Activity(mCi)",
                   "Permitted Date", "Recommended Date", "Fraction", "Limit(kBq/kg)", "Bag Mass(kg)"]
    for sheet_name in ("Vials",) :
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(vial_headers)
    for sheet_name in ("Tc99m", "Ga68", "Lu177"):
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Batch / Item Log"])
    wb.save(xlsx_path)
    return wb

def ensure_daily_log_sqlite():
    ensure_dir(DAILY_DISPOSALS_DIR)
    conn  = sqlite3.connect(DAILY_LOG_DB)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS disposed_vials (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                              disposal_date TEXT NOT NULL,
                                                              disposal_time TEXT NOT NULL,
                                                              radionuclide TEXT NOT NULL,
                                                              calibration_date TEXT,
                                                              stored_at TEXT NOT NULL,
                                                              activity_mci REAL NOT NULL,
                                                              permitted_date TEXT,
                                                              recommended_date TEXT,
                                                              limit_kbq_kg REAL,
                                                              fraction REAL,
                                                              bag_mass REAL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS disposed_tc99m_batches (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                                      batch_id INTEGER NOT NULL,
                                                                      item_id TEXT NOT NULL,
                                                                      stored_at TEXT NOT NULL,
                                                                      activity_mci REAL NOT NULL,
                                                                      permitted_date TEXT,
                                                                      recommended_date TEXT,
                                                                      limit_kbq_kg REAL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS disposed_ga68_batches (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                                     batch_id INTEGER NOT NULL,
                                                                     item_id TEXT NOT NULL,
                                                                     stored_at TEXT NOT NULL,
                                                                     activity_mci REAL NOT NULL,
                                                                     permitted_date TEXT,
                                                                     recommended_date TEXT,
                                                                     limit_kbq_kg REAL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS disposed_lu177_batches (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                                      batch_id INTEGER NOT NULL,
                                                                      item_id TEXT NOT NULL,
                                                                      stored_at TEXT NOT NULL,
                                                                      activity_mci REAL NOT NULL,
                                                                      permitted_date TEXT,
                                                                      recommended_date TEXT,
                                                                      limit_kbq_kg REAL)""")
    cols = [r[1] for r in cur.execute("PRAGMA table_info(disposed_vials)").fetchall()]
    if "limit_kbq_kg" not in cols:
        cur.execute("ALTER TABLE disposed_vials ADD COLUMN limit_kbq_kg REAL")
    if "bag_mass" not in cols:
        cur.execute("ALTER TABLE disposed_vials ADD COLUMN bag_mass REAL")
    if "fraction" not in cols:
        cur.execute("ALTER TABLE disposed_vials ADD COLUMN fraction REAL")
    for table in ("disposed_tc99m_batches", "disposed_ga68_batches", "disposed_lu177_batches"):
        cols = [r[1] for r in cur.execute(f"PRAGMA table_info({table})").fetchall()]
        if "limit_kbq_kg" not in cols:
            cur.execute(f"ALTER TABLE {table} ADD COLUMN limit_kbq_kg REAL")
    conn.commit()
    return conn

def log_vials_disposal(vials_full_rows, *, bag_mass_kg=None, fraction=None):
    disp_date = datetime.now().strftime(DATE_FORMAT)
    disp_time = datetime.now().strftime(HOUR_FORMAT)
    xlsx_path = get_daily_disposal_excel_path(disp_date)
    wb = ensure_daily_log_workbook(xlsx_path)
    ws = wb["Vials"]
    conn = ensure_daily_log_sqlite()
    cur = conn.cursor()
    for r in vials_full_rows:
        (rid, radionuclide, source_db, calibration_date, stored_at, activity_mci, permitted_date, recommended_date, limit_bq, limit_mci) = r
        limit_kbq_kg = DISPOSAL_LIMITS_KBQ_PER_KG.get(radionuclide)
        ws.append([disp_date, disp_time, radionuclide, calibration_date, stored_at, float(activity_mci), permitted_date, recommended_date,
                   limit_kbq_kg, fraction, bag_mass_kg])
        cur.execute("INSERT INTO disposed_vials (disposal_date, disposal_time, radionuclide, calibration_date, stored_at, activity_mci, permitted_date, recommended_date, limit_kbq_kg, fraction, bag_mass) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?)", (disp_date, disp_time, radionuclide, calibration_date, stored_at, float(activity_mci), permitted_date, recommended_date,
                                                      limit_kbq_kg, fraction, bag_mass_kg))
        mark_vial_as_disposed(source_db=source_db, disposed_date=disp_date)
    wb.save(xlsx_path)
    conn.commit()
    conn.close()

def append_batch(ws, *, batch_name, finalized_at, disposed_dt_str, items_rows, radionuclide):
    conn = ensure_daily_log_sqlite()
    cur = conn.cursor()
    start_row = ws.max_row + 1
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        start_row = 1
    limit_txt = DISPOSAL_LIMITS_KBQ_PER_KG.get(radionuclide, "-")
    title = f"BATCH: {batch_name} | Finalized: {finalized_at or '-'} | Disposed: {disposed_dt_str}"
    ws.cell(row=start_row, column=1, value=title)
    ws.cell(row=start_row, column=1).font = Font("bold")
    ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
    ws.cell(row=start_row + 1, column=1, value=f"(radionuclide={radionuclide}, limit(kBq/kg)={limit_txt})")
    ws.cell(row=start_row + 1, column=1).alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=9)
    headers = ["ID", "Stored At", "Activity(mCi)", "Permitted Date", "Recommended Date"]
    header_row = start_row + 2
    for c,h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font("bold")
        cell.alignment = Alignment(horizontal="center")
    r = header_row + 1
    for item in items_rows:
        iid, item_label, stored_at, activity_mci, permitted, recommended = item
        ws.cell(row=r, column=1, value=item_label)
        ws.cell(row=r, column=2, value=stored_at)
        ws.cell(row=r, column=3, value=float(activity_mci))
        ws.cell(row=r, column=4, value=permitted)
        ws.cell(row=r, column=5, value=recommended)
        r += 1
    ws.append([])
    widths = [12, 18, 18, 18, 18, 8]
    for i,w in enumerate(widths, start=1):
        col = get_column_letter(i)
        cw = ws.column_dimensions[col].width
        if cw is None or cw < w:
            ws.column_dimensions[col].width = w
    for item in items_rows:
        iid, item_label, stored_at, activity_mci, permitted, recommended = item
        cur.execute(f"INSERT INTO disposed_{radionuclide.lower()}_batches (batch_id, item_id, stored_at, activity_mci, permitted_date, recommended_date, limit_kbq_kg)"
                    "VALUES (?,?,?,?,?,?,?)", (batch_name, iid, stored_at, float(activity_mci), permitted, recommended, None if limit_txt in (None,"-") else float(limit_txt)))
    conn.commit()
    conn.close()

def log_batch_disposal(batch_path: str, finalized_at: str, items_rows, *, radionuclide):
    disp_date = datetime.now().strftime(DATE_FORMAT)
    disp_time = datetime.now().strftime(HOUR_FORMAT)
    disp_dt_str = f"{disp_date} {disp_time}"
    batch_name = os.path.basename(batch_path)
    xlsx_path = get_daily_disposal_excel_path(disp_date)
    wb = ensure_daily_log_workbook(xlsx_path)
    ws = wb[f"{radionuclide}"]
    append_batch(ws, batch_name=batch_name, finalized_at=finalized_at, disposed_dt_str=disp_dt_str, items_rows=items_rows, radionuclide=radionuclide)
    wb.save(xlsx_path)

#=====MARK VIAL AS DISPOSED=====
def mark_vial_as_disposed(source_db, disposed_date):
    conn = sqlite3.connect(source_db)
    cur = conn.cursor()
    cur.execute("UPDATE vial_info SET disposal_date=?", (disposed_date,))
    conn.commit()
    conn.close()
    source_excel = os.path.splitext(source_db)[0] + ".xlsx"
    if os.path.exists(source_excel):
        wb = load_workbook(source_excel)
        ws = wb["Vial Info"]
        if ws.max_row >= 2:
            ws.cell(row=2, column=8, value=disposed_date)
        wb.save(source_excel)

#=====VIALS SQLITE + (LIVE) EXCEL=====
def init_vials_storage():
    ensure_dir(VIALS_STORAGE_DIR)
    conn = sqlite3.connect(VIALS_DB)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS stored_vials (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                                            radionuclide TEXT,
                                                            calibration_date TEXT,
                                                            stored_at TEXT,
                                                            activity_mci REAL,
                                                            permitted_date TEXT,
                                                            recommended_date TEXT,
                                                            limit_mci REAL,
                                                            limit_bq REAL,
                                                            source_db TEXT)""")
    conn.commit()
    conn.close()
    if not os.path.exists(VIALS_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stored Vials"
        ws.append(["ID", "Radionuclide", "Calibration Date", "Stored At", "Activity(mCi)", "Permitted Date",
                   "Recommended Date", "Limit(mCi)", "Limit(Bq)", "Source DB"])
        wb.save(VIALS_XLSX)

#=====STORE VIAL IN SQLITE + EXCEL=====
def store_vial(radionuclide, source_db, calibration_date, stored_at, activity_mci, permitted_date=None, recommended_date=None, limit_bq=None):
    if permitted_date is None or recommended_date is None:
        recommended_date, permitted_date, limit_bq = calc_recommended_and_permitted_date(radionuclide, float(activity_mci), stored_at)
    if recommended_date is None or permitted_date is None:
        return None, None
    limit_mci = None
    if limit_bq not in (None, 0):
        limit_mci = round(bq_to_mci(limit_bq), 2)
    conn = sqlite3.connect(VIALS_DB)
    cur = conn.cursor()
    cur.execute("INSERT INTO stored_vials (radionuclide, calibration_date, stored_at, activity_mci, permitted_date, recommended_date, limit_mci, limit_bq, source_db) VALUES (?,?,?,?,?,?,?,?,?)",
                (radionuclide, calibration_date, stored_at, float(activity_mci), permitted_date, recommended_date, None if limit_mci is None else float(limit_mci), None if limit_bq is None else float(limit_bq), source_db))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    wb = load_workbook(VIALS_XLSX)
    ws = wb["Stored Vials"]
    ws.append([new_id, radionuclide, calibration_date, stored_at, float(activity_mci), permitted_date, recommended_date, "" if limit_mci is None else float(limit_mci), "" if limit_bq is None else float(limit_bq), source_db])
    wb.save(VIALS_XLSX)
    return new_id

#=====READ STORED VIALS FROM BATCH=====
def read_stored_vials():
    init_vials_storage()
    conn = sqlite3.connect(VIALS_DB)
    cur = conn.cursor()
    rows = cur.execute("SELECT id, radionuclide, calibration_date, stored_at, activity_mci, permitted_date, recommended_date, limit_mci, source_db FROM stored_vials ORDER BY id").fetchall()
    conn.close()
    return rows

#=====READ VIALS IDS=====
def read_vials_full_ids(ids):
    if not ids:
        return []
    init_vials_storage()
    conn = sqlite3.connect(VIALS_DB)
    cur = conn.cursor()
    placeholders = ",".join(["?"] * len(ids))
    rows = cur.execute(f"SELECT id, radionuclide, source_db, calibration_date, stored_at, activity_mci, permitted_date, recommended_date, limit_bq, limit_mci FROM stored_vials WHERE id IN ({placeholders}) ORDER BY id", ids).fetchall()
    conn.close()
    return rows

def get_vial_cal_activity(source_db):
    try:
        conn = sqlite3.connect(source_db)
        cur = conn.cursor()
        row = cur.execute("SELECT activity FROM vial_info ORDER BY rowid DESC LIMIT 1").fetchone()
        conn.close()
        if row and row[0] is not None:
            return float(row[0])
    except Exception:
        pass
    return None

#=====DELETE VIALS BY IDS=====
def delete_vials_by_ids(ids):
    if not ids:
        return
    init_vials_storage()
    conn = sqlite3.connect(VIALS_DB)
    cur = conn.cursor()
    placeholder = ",".join(["?"] * len(ids))
    cur.execute(f"DELETE FROM stored_vials WHERE id IN ({placeholder})", ids)
    conn.commit()
    conn.close()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Stored Vials"
        ws.append(["ID", "Radionuclide", "Calibration Date", "Stored At", "Activity(mCi)", "Permitted Date",
                    "Recommended Date", "Limit(mCi)", "Limit(Bq)", "Source DB"])
        conn2 = sqlite3.connect(VIALS_DB)
        cur2 = conn2.cursor()
        rows = cur2.execute("SELECT id, radionuclide, calibration_date, stored_at, activity_mci, permitted_date, recommended_date, limit_mci, limit_bq, source_db FROM stored_vials ORDER BY id").fetchall()
        conn2.close()
        for r in rows:
            ws.append(list(r))
        wb.save(VIALS_XLSX)
    except Exception:
        pass

#=====GROUP READY VIALS=====
def calc_bag_clearance_smart(ready_items, mass_kg):
    mass_kg = float(mass_kg)
    enriched = []
    for it in ready_items:
        nuclide = it["radionuclide"]
        limit = DISPOSAL_LIMITS_KBQ_PER_KG[nuclide]
        frac = (mci_to_kbq(it["activity_now"]) / mass_kg) / float(limit)
        enriched.append((frac,it))
    enriched.sort(key=lambda x: x[0])
    selected = []
    skipped = []
    total_fraction = 0.0
    for frac, it in enriched:
        if total_fraction + frac < 1.0:
            selected.append(it)
            total_fraction += frac
        else:
            skipped.append(it)
    details, _, _ = calc_bag_clearance(selected, mass_kg)
    return details, total_fraction, selected, skipped

#=====TC99M SQLITE CREATE TABLES=====
def init_registry(base_dir, registry_db):
    ensure_dir(base_dir)
    conn = sqlite3.connect(registry_db)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY, value TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS batches (id INTEGER PRIMARY KEY AUTOINCREMENT, folder_path TEXT, created_at TEXT, finalized_at TEXT, disposed_at TEXT)""")
    cols = [r[1] for r in cur.execute("PRAGMA table_info(batches)").fetchall()]
    if "disposed_at" not in cols:
        cur.execute("ALTER TABLE batches ADD COLUMN disposed_at TEXT")
    if "finalized_at" not in cols:
        cur.execute("ALTER TABLE batches ADD COLUMN finalized_at TEXT")
    conn.commit()
    conn.close()

#=====CREATE NEW BATCH=====
def create_new_batch_folder(base_dir, registry_db):
    init_registry(base_dir, registry_db)
    year = datetime.now().strftime("%Y")
    creation_date = datetime.now().strftime(DATE_FORMAT)
    year_dir = os.path.join(base_dir, year)
    ensure_dir(year_dir)
    batch_folder = f"Batch__{creation_date}-ACTIVE"
    batch_path = os.path.join(year_dir, batch_folder)
    if not os.path.exists(batch_path):
        ensure_dir(batch_path)
    conn = sqlite3.connect(registry_db)
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM batches WHERE folder_path=?", (batch_path,))
    exists = cur.fetchone()
    if not exists:
        cur.execute("INSERT INTO batches (folder_path, created_at, finalized_at) VALUES (?,?,NULL)", (batch_path, creation_date))
    cur.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('active_batch',?)", (batch_path,))
    conn.commit()
    conn.close()
    return batch_path

#=====GET ACTIVE BATCH (IF IT DOESN'T EXIST CREATE NEW)=====
def get_active_batch(base_dir, registry_db):
    init_registry(base_dir, registry_db)
    conn = sqlite3.connect(registry_db)
    cur = conn.cursor()
    row = cur.execute("SELECT value FROM settings WHERE key='active_batch'").fetchone()
    conn.close()
    if row and os.path.isdir(row[0]):
        return row[0]
    return create_new_batch_folder(base_dir, registry_db)

#=====FINALIZE CURRENT ACTIVE BATCH AND START NEW=====
def finalize_active_batch(base_dir, registry_db):
    init_registry(base_dir, registry_db)
    old_batch = get_active_batch(base_dir, registry_db)
    finalized_date = mark_items_finalized(old_batch)
    update_storage_excel_batch_column(old_batch, 7, finalized_date)
    stored_batch = update_folder_status(old_batch, stored=True)
    conn = sqlite3.connect(registry_db)
    cur = conn.cursor()
    cur.execute("UPDATE batches SET finalized_at=?, folder_path=? WHERE folder_path=?", (finalized_date, stored_batch, old_batch))
    conn.commit()
    conn.close()
    new_batch = create_new_batch_folder(base_dir, registry_db)
    return stored_batch, new_batch

#=====DISPOSE BATCH=====
def dispose_batch(batch_path, base_dir, registry_db):
    init_registry(base_dir, registry_db)
    disposed_date = mark_items_disposed(batch_path)
    update_storage_excel_batch_column(batch_path, 8, disposed_date)
    new_path = update_folder_status(batch_path, disposed=True)
    conn = sqlite3.connect(registry_db)
    cur = conn.cursor()
    cur.execute("UPDATE batches SET disposed_at=?, folder_path=? WHERE folder_path=?", (disposed_date, new_path, batch_path))
    conn.commit()
    conn.close()
    return new_path

#=====READ BATCH DATE INFO=====
def read_batch_info(batch_path, base_dir=None, registry_db=None):
    init_storage_files(batch_path)
    db_path = os.path.join(batch_path, "storage.sqlite")
    if not os.path.exists(db_path):
        return None, None, None
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    row = cur.execute("SELECT MIN(batch_created_at), MAX(batch_finalized_at), MAX(batch_disposed_at) FROM stored_items").fetchone()
    conn.close()
    if not row:
        return None, None, None
    return row[0], row[1], row[2]

#=====CREATE SQLITE+XLSX FILES INSIDE BATCH FOLDER=====
def init_storage_files(batch_path):
    db_path = os.path.join(batch_path, "storage.sqlite")
    xlsx_path = os.path.join(batch_path, "storage.xlsx")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS stored_items (id TEXT PRIMARY KEY,
                                                            source_parent_id TEXT,
                                                            item_label TEXT,
                                                            kit_name TEXT,
                                                            stored_at TEXT,
                                                            activity_mci REAL,
                                                            permitted_date TEXT,
                                                            recommended_date TEXT,
                                                            batch_created_at TEXT,
                                                            batch_finalized_at TEXT,
                                                            batch_disposed_at TEXT)""")
    cols = [r[1] for r in cur.execute("PRAGMA table_info(stored_items)").fetchall()]
    for col in ["batch_created_at", "batch_finalized_at", "batch_disposed_at"]:
        if col not in cols:
            cur.execute(f"ALTER TABLE stored_items ADD COLUMN {col} TEXT")
    conn.commit()
    conn.close()
    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stored Items"
        ws.append(["ID", "Stored at", "Activity(mCi)", "Permitted Date", "Recommended Date", "Batch Created At", "Batch Finalized At", "Batch Disposed At"])
        wb.save(xlsx_path)
    return db_path, xlsx_path

def update_storage_excel_batch_column(batch_path, column, value):
    xlsx_path = os.path.join(batch_path, "storage.xlsx")
    wb = load_workbook(xlsx_path)
    ws = wb["Stored Items"]
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=column, value=value)
    wb.save(xlsx_path)

#=====STORE TC99M ITEM IN SQLITE+XLSX=====
def store_item(item_id, source_parent_id, item_label, kit_name, stored_at, activity_mci, *, radionuclide, base_dir, registry_db, permitted_date=None, recommended_date=None):
    batch_path = get_active_batch(base_dir, registry_db)
    db_path, xlsx_path = init_storage_files(batch_path)
    created_at = datetime.now().strftime(DATE_FORMAT)
    if permitted_date is None or recommended_date is None:
        recommended_date, permitted_date, _ = calc_recommended_and_permitted_date(radionuclide=radionuclide, activity_mci=float(activity_mci), stored_at=stored_at)
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("INSERT OR IGNORE INTO stored_items (id, source_parent_id, item_label, kit_name, stored_at, activity_mci, permitted_date, recommended_date, batch_created_at, batch_finalized_at, batch_disposed_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (item_id, str(source_parent_id), item_label, kit_name, stored_at, float(activity_mci), permitted_date, recommended_date, created_at, None, None))
    inserted = cur.rowcount
    conn.commit()
    conn.close()
    if inserted:
        wb = load_workbook(xlsx_path)
        ws = wb["Stored Items"]
        ws.append([item_id, stored_at, float(activity_mci), permitted_date, recommended_date, created_at, None, None])
        wb.save(xlsx_path)
    return item_id, batch_path

#=====READ ITEMS INFO=====
def read_items(batch_path=None, *, base_dir, registry_db):
    if batch_path is None:
        batch_path = get_active_batch(base_dir, registry_db)
    db_path = os.path.join(batch_path, "storage.sqlite")
    if not os.path.exists(db_path):
        return []
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    rows = cur.execute("SELECT id, item_label, stored_at, activity_mci, permitted_date, recommended_date FROM stored_items ORDER BY stored_at, id").fetchall()
    conn.close()
    return rows

def mark_items_finalized(batch_path):
    finalized_date = datetime.now().strftime(DATE_FORMAT)
    db_path = os.path.join(batch_path, "storage.sqlite")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("UPDATE stored_items SET batch_finalized_at=?", (finalized_date,))
    conn.commit()
    conn.close()
    return finalized_date

def mark_items_disposed(batch_path):
    disposed_date = datetime.now().strftime(DATE_FORMAT)
    db_path = os.path.join(batch_path, "storage.sqlite")
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("UPDATE stored_items SET batch_disposed_at=?", (disposed_date,))
    conn.commit()
    conn.close()
    return disposed_date

#====SYNC TC99M+GA68 GEN FOR DISPOSAL=====
def ensure_elutions_disposal_columns(cur, conn):
    cols = [r[1] for r in cur.execute("PRAGMA table_info(elutions)").fetchall()]
    if "stored_to_disposal" not in cols:
        cur.execute("ALTER TABLE elutions ADD COLUMN stored_to_disposal INTEGER DEFAULT 0")
    if "stored_to_disposal_at" not in cols:
        cur.execute("ALTER TABLE elutions ADD COLUMN stored_to_disposal_at TEXT")
    if "disposal_item_id" not in cols:
        cur.execute("ALTER TABLE elutions ADD COLUMN disposal_item_id TEXT")
    conn.commit()

def ensure_kits_disposal_columns(cur, conn):
    cols = [r[1] for r in cur.execute("PRAGMA table_info(kits)").fetchall()]
    if "stored_to_disposal" not in cols:
        cur.execute("ALTER TABLE kits ADD COLUMN stored_to_disposal INTEGER DEFAULT 0")
    if "stored_to_disposal_at" not in cols:
        cur.execute("ALTER TABLE kits ADD COLUMN stored_to_disposal_at TEXT")
    if "disposal_item_id" not in cols:
        cur.execute("ALTER TABLE kits ADD COLUMN disposal_item_id TEXT")
    conn.commit()

def sync_tc99m_elutions_for_disposal(dbfile):
    conn = sqlite3.connect(dbfile)
    cur = conn.cursor()
    ensure_elutions_disposal_columns(cur, conn)
    today = datetime.now().date()
    elution_rows = cur.execute("SELECT id, date, time, volume, concentration "
                               "FROM elutions "
                               "WHERE COALESCE(stored_to_disposal, 0) = 0 "
                               "ORDER BY date, time").fetchall()
    synced_count = 0
    for elution_id, date_str, time_str, volume, concentration in elution_rows:
        try:
            elution_date = datetime.strptime(date_str, DATE_FORMAT).date()
        except ValueError:
            continue
        if elution_date >= today:
            continue
        volume = float(volume or 0)
        concentration = float(concentration or 0)
        kit_rows = cur.execute("SELECT time, volume FROM kits WHERE parent_id IS NULL AND date=? AND elution=? ORDER BY time",
                               (date_str, time_str)).fetchall()
        used_volume = 0.0
        last_labeling_time = time_str
        for kit_time, kit_volume in kit_rows:
            used_volume += float(kit_volume or 0)
            if kit_time:
                last_labeling_time = kit_time
        remaining_volume = round(volume - used_volume, 2)
        if remaining_volume < 0:
            remaining_volume = 0.0
        try:
            elution_dt = datetime.strptime(f"{date_str} {time_str}", f"{DATE_FORMAT} {HOUR_FORMAT}")
            last_labeling_dt = datetime.strptime(f"{date_str} {last_labeling_time}", f"{DATE_FORMAT} {HOUR_FORMAT}")
            delta_hours = (last_labeling_dt - elution_dt).total_seconds() / 3600.0
            if delta_hours < 0:
                delta_hours = 0.0
            decayed_conc = decay_activity(concentration, T12_TC99M, delta_hours)
        except Exception:
            decayed_conc = concentration
        remaining_activity = round(remaining_volume * decayed_conc, 6)
        stored_at = date_str
        item_id = f"Elution-{time_str}"
        item_label = f"Elution-{time_str}"
        if remaining_activity > 0:
            recommended_date, permitted_date, _ = calc_recommended_and_permitted_date(radionuclide=TC99M_NUCLIDE,
                                                                                      activity_mci=remaining_activity,
                                                                                      stored_at=stored_at)
            store_item(item_id=item_id,
                       source_parent_id=f"ELUTION-{elution_id}",
                       item_label=item_label,
                       kit_name="ELUTION",
                       stored_at=stored_at,
                       activity_mci=remaining_activity,
                       radionuclide=TC99M_NUCLIDE,
                       base_dir=TC99M_DIR,
                       registry_db=TC99M_REGISTRY_DB,
                       permitted_date=permitted_date,
                       recommended_date=recommended_date)
            cur.execute("UPDATE elutions SET stored_to_disposal = 1, stored_to_disposal_at = ?, disposal_item_id = ? WHERE id=?",
                        (datetime.now().strftime(DATE_FORMAT), item_id, elution_id))
            synced_count += 1
    conn.commit()
    conn.close()
    return synced_count

def sync_tc99m_kits_for_disposal(dbfile):
    conn = sqlite3.connect(dbfile)
    cur = conn.cursor()
    ensure_kits_disposal_columns(cur, conn)
    today = datetime.now().date()
    rows = cur.execute("SELECT id, date, time, kit, concentration, volume_left FROM kits WHERE parent_id IS NULL AND COALESCE(stored_to_disposal, 0) = 0 ORDER BY date, time").fetchall()
    synced_count = 0
    for parent_id, kit_date_str, kit_time_str, kit_name, concentration, volume_left in rows:
        try:
            kit_date = datetime.strptime(kit_date_str, DATE_FORMAT).date()
        except ValueError:
            continue
        if kit_date >= today:
            continue
        parent_concentration = float(concentration or 0)
        parent_volume_left = float(volume_left or 0)
        child_rows = cur.execute("SELECT time, dose_volume FROM kits WHERE parent_id=? ORDER BY time",
                                 (parent_id,)).fetchall()
        used_volume = 0.0
        last_child_time = kit_time_str
        for child_time, child_dose_volume in child_rows:
            used_volume += float(child_dose_volume or 0)
            if child_time:
                last_child_time = child_time
        remaining_volume = round(parent_volume_left - used_volume, 2)
        if remaining_volume < 0:
            remaining_volume = 0.0
        try:
            parent_dt = datetime.strptime(f"{kit_date_str} {kit_time_str}", f"{DATE_FORMAT} {HOUR_FORMAT}")
            last_child_dt = datetime.strptime(f"{kit_date_str} {last_child_time}", f"{DATE_FORMAT} {HOUR_FORMAT}")
            delta_hours = (last_child_dt - parent_dt).total_seconds() / 3600
            if delta_hours < 0:
                delta_hours = 0.0
            decayed_conc = decay_activity(parent_concentration, T12_TC99M, delta_hours)
        except Exception:
            decayed_conc = parent_concentration
        residual_activity = round(remaining_volume * decayed_conc, 6)
        stored_at = kit_date_str
        item_id = f"{kit_name}-{kit_time_str}"
        item_label = f"{kit_name}-{kit_time_str}"
        if residual_activity > 0:
            recommended_date, permitted_date, _ = calc_recommended_and_permitted_date(radionuclide=TC99M_NUCLIDE,
                                                                                      activity_mci=residual_activity,
                                                                                      stored_at=stored_at)
            store_item(item_id=item_id,
                       source_parent_id=parent_id,
                       item_label=item_label,
                       kit_name=kit_name,
                       stored_at=stored_at,
                       activity_mci=residual_activity,
                       radionuclide=TC99M_NUCLIDE,
                       base_dir=TC99M_DIR,
                       registry_db=TC99M_REGISTRY_DB,
                       permitted_date=permitted_date,
                       recommended_date=recommended_date)
        cur.execute("UPDATE kits SET stored_to_disposal = 1, stored_to_disposal_at = ?, disposal_item_id = ? WHERE id=?",
                    (datetime.now().strftime(DATE_FORMAT), item_id, parent_id))
        synced_count += 1
    conn.commit()
    conn.close()
    return synced_count

def sync_tc99m_gen_for_disposal(dbfile):
    kits_count = sync_tc99m_kits_for_disposal(dbfile)
    elutions_count = sync_tc99m_elutions_for_disposal(dbfile)
    return kits_count, elutions_count

def ensure_dotatoc_disposal_columns(cur, conn):
    cols = [r[1] for r in cur.execute("PRAGMA table_info(dotatoc)").fetchall()]
    if "stored_to_disposal" not in cols:
        cur.execute("ALTER TABLE dotatoc ADD COLUMN stored_to_disposal INTEGER DEFAULT 0")
    if "stored_to_disposal_at" not in cols:
        cur.execute("ALTER TABLE dotatoc ADD COLUMN stored_to_disposal_at TEXT")
    if "disposal_item_id" not in cols:
        cur.execute("ALTER TABLE dotatoc ADD COLUMN disposal_item_id TEXT")
    conn.commit()

def sync_ga68_elutions_for_disposal(dbfile):
    conn = sqlite3.connect(dbfile)
    cur = conn.cursor()
    ensure_elutions_disposal_columns(cur, conn)
    today = datetime.now().date()
    elution_rows = cur.execute(
        "SELECT id, date, time, activity "
        "FROM elutions "
        "WHERE COALESCE(stored_to_disposal, 0) = 0 "
        "ORDER BY date, time").fetchall()
    synced_count = 0
    for elution_id, date_str, time_str, activity in elution_rows:
        try:
            elution_date = datetime.strptime(date_str, DATE_FORMAT).date()
        except ValueError:
            continue
        if elution_date >= today:
            continue
        initial_activity = float(activity or 0)
        dotatoc_rows = cur.execute(
            "SELECT admin_time, dose "
            "FROM dotatoc "
            "WHERE date=? "
            "ORDER BY admin_time",
            (date_str,)).fetchall()
        used_activity = 0.0
        last_use_time = time_str
        for admin_time, dose in dotatoc_rows:
            used_activity += float(dose or 0)
            if admin_time:
                last_use_time = admin_time
        remaining_activity_at_elution = initial_activity - used_activity
        if remaining_activity_at_elution < 0:
            remaining_activity_at_elution = 0.0
        try:
            elution_dt = datetime.strptime(f"{date_str} {time_str}", f"{DATE_FORMAT} {HOUR_FORMAT}")
            last_use_dt = datetime.strptime(f"{date_str} {last_use_time}", f"{DATE_FORMAT} {HOUR_FORMAT}")
            delta_minutes = (last_use_dt - elution_dt).total_seconds() / 60.0
            if delta_minutes < 0:
                delta_minutes = 0.0
            decay_factor = math.exp(-(math.log(2) / T12_GA68) * delta_minutes)
            remaining_activity = round(remaining_activity_at_elution * decay_factor, 6)
        except Exception:
            remaining_activity = round(remaining_activity_at_elution, 6)
        stored_at = date_str
        item_id = f"Ga68-Elution-{time_str}"
        item_label = f"Elution-{time_str}"
        if remaining_activity > 0:
            recommended_date, permitted_date, _ = calc_recommended_and_permitted_date(radionuclide="Ga68",
                                                                                      activity_mci=remaining_activity,
                                                                                      stored_at=stored_at)
            store_item(item_id=item_id,
                       source_parent_id=f"GA68-ELUTION-{elution_id}",
                       item_label=item_label,
                       kit_name="ELUTION",
                       stored_at=stored_at,
                       activity_mci=remaining_activity,
                       radionuclide="Ga68",
                       base_dir=GA68_DIR,
                       registry_db=GA68_REGISTRY_DB,
                       permitted_date=permitted_date,
                       recommended_date=recommended_date)
        cur.execute("UPDATE elutions SET stored_to_disposal=1, stored_to_disposal_at=?, disposal_item_id=? WHERE id=?",
                    (datetime.now().strftime(DATE_FORMAT), item_id, elution_id))
        synced_count += 1
    conn.commit()
    conn.close()
    return synced_count

def sync_ga68_dotatoc_for_disposal(dbfile):
    conn = sqlite3.connect(dbfile)
    cur = conn.cursor()
    ensure_dotatoc_disposal_columns(cur, conn)
    today = datetime.now().date()
    rows = cur.execute(
        "SELECT id, date, patient, admin_time, residual "
        "FROM dotatoc "
        "WHERE COALESCE(stored_to_disposal, 0) = 0 "
        "ORDER BY date, admin_time").fetchall()
    synced_count = 0
    for row_id, date_str, patient, admin_time, residual in rows:
        try:
            d = datetime.strptime(date_str, DATE_FORMAT).date()
        except ValueError:
            continue
        if d >= today:
            continue
        residual_activity = float(residual or 0)
        item_id = f"Ga68-DOTATOC-{row_id}"
        item_label = f"DOTATOC-{admin_time}"
        if residual_activity > 0:
            recommended_date, permitted_date, _ = calc_recommended_and_permitted_date(radionuclide="Ga68",
                                                                                      activity_mci=residual_activity,
                                                                                      stored_at=date_str)
            store_item(item_id=item_id,
                       source_parent_id=f"DOTATOC-{row_id}",
                       item_label=item_label,
                       kit_name="DOTATOC",
                       stored_at=date_str,
                       activity_mci=residual_activity,
                       radionuclide="Ga68",
                       base_dir=GA68_DIR,
                       registry_db=GA68_REGISTRY_DB,
                       permitted_date=permitted_date,
                       recommended_date=recommended_date)
        cur.execute("UPDATE dotatoc SET stored_to_disposal=1, stored_to_disposal_at=?, disposal_item_id=? WHERE id=?",
                    (datetime.now().strftime(DATE_FORMAT), item_id, row_id))
        synced_count += 1
    conn.commit()
    conn.close()
    return synced_count

def sync_ga68_gen_for_disposal(dbfile):
    elutions_count = sync_ga68_elutions_for_disposal(dbfile)
    dotatoc_count = sync_ga68_dotatoc_for_disposal(dbfile)
    return dotatoc_count, elutions_count
